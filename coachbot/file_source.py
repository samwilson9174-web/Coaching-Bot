"""
file_source.py — FILE-BASED DATA INGESTION (manual Excel/CSV export)
--------------------------------------------------------------------
For when you DON'T have live MT5 Manager API access. You export the trade
history manually from MT5 / Brokeret as .csv or .xlsx, drop it in a folder, and
the bot reads it. Same normalised output shape as the MT5 source, so everything
downstream is identical.

YOU PROVIDE TWO FILES (paths set in .env):
  TRADES_FILE  : the trade-history export (one row per closed trade)
  USERS_FILE   : client list + consent + telegram handle

COLUMN MAPPING
--------------
Exports vary. Instead of forcing exact headers, we map common aliases to our
canonical fields (case-insensitive, ignores spaces/underscores). If your export
uses a name we don't recognise, add it to the alias lists below — that's the
only place you'd edit.

Canonical trade fields:
  login, ticket, symbol, type, volume, open_price, close_price,
  sl, tp, sl_hit, tp_hit, open_time, close_time, profit

Canonical user fields:
  login, name, email, telegram_id, consent
"""
from __future__ import annotations
import os
import csv
from datetime import datetime

from .logger import get_logger

log = get_logger("file")


# ---- column aliases (lowercase, alphanumeric only) ------------------------
TRADE_ALIASES = {
    "login":       ["login", "account", "accountid", "clientlogin", "userlogin", "mt5login"],
    "ticket":      ["ticket", "order", "orderid", "dealid", "positionid", "deal", "position"],
    "symbol":      ["symbol", "instrument", "pair", "asset"],
    "type":        ["type", "direction", "side", "action", "buysell", "ordertype"],
    "volume":      ["volume", "lots", "lot", "size", "qty", "quantity"],
    "open_price":  ["openprice", "priceopen", "entryprice", "entry", "open"],
    "close_price": ["closeprice", "priceclose", "exitprice", "exit", "close"],
    "sl":          ["sl", "stoploss", "stop", "slprice", "pricesl"],
    "tp":          ["tp", "takeprofit", "target", "tpprice", "pricetp"],
    "sl_hit":      ["slhit", "stoplosshit", "hitsl", "wasslhit"],
    "tp_hit":      ["tphit", "takeprofithit", "hittp", "wastphit"],
    "open_time":   ["opentime", "timeopen", "opendate", "entrytime", "openat"],
    "close_time":  ["closetime", "timeclose", "closedate", "exittime", "closeat"],
    "profit":      ["profit", "pnl", "pl", "netprofit", "result", "profitloss"],
}

USER_ALIASES = {
    "login":       ["login", "account", "accountid", "clientlogin", "userlogin", "mt5login"],
    "name":        ["name", "clientname", "fullname", "client", "customer"],
    "email":       ["email", "emailaddress", "mail", "e-mail"],
    "telegram_id": ["telegramid", "telegram", "chatid", "telegramchatid", "tg", "telegramusername"],
    "consent":     ["consent", "consentoptin", "optin", "subscribed", "marketingconsent"],
}


def _norm(s):
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


def _build_index(header, aliases):
    """Map canonical field -> actual column index in this file."""
    norm_header = [_norm(h) for h in header]
    index = {}
    for canon, names in aliases.items():
        for name in names:
            if name in norm_header:
                index[canon] = norm_header.index(name)
                break
    return index


def _read_rows(path):
    """Return (header, rows) from a .csv or .xlsx file."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl required for Excel files: pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [[("" if c is None else c) for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        if not rows:
            return [], []
        return rows[0], rows[1:]
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = list(csv.reader(f))
        if not reader:
            return [], []
        return reader[0], reader[1:]


# ---- value normalisers ----------------------------------------------------
def _to_float(v, default=0.0):
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return default


def _to_direction(v):
    s = str(v).strip().lower()
    if s in ("buy", "long", "0", "buy limit", "buy stop"):
        return "buy"
    if s in ("sell", "short", "1", "sell limit", "sell stop"):
        return "sell"
    return "buy" if "buy" in s or "long" in s else "sell"


def _to_bool_int(v):
    s = str(v).strip().lower()
    return 1 if s in ("1", "true", "yes", "y", "hit") else 0


def _to_time(v):
    """Accept several common datetime formats; return 'YYYY-MM-DD HH:MM:SS'."""
    s = str(v).strip()
    if not s:
        return ""
    # Already a datetime (Excel cells)
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    fmts = ["%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M:%S", "%Y-%m-%d %H:%M",
            "%Y.%m.%d %H:%M", "%d/%m/%Y %H:%M:%S", "%m/%d/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M", "%Y-%m-%dT%H:%M:%S"]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    # last resort: take first 19 chars if it looks like a timestamp
    return s[:19]


class FileDataSource:
    def __init__(self, config):
        self.config = config
        self._users = []
        self._trades_by_login = {}

    def connect(self):
        trades_file = self.config.TRADES_FILE
        users_file = self.config.USERS_FILE
        if not trades_file or not os.path.exists(trades_file):
            raise FileNotFoundError(f"TRADES_FILE not found: {trades_file}")
        if not users_file or not os.path.exists(users_file):
            raise FileNotFoundError(f"USERS_FILE not found: {users_file}")

        self._load_users(users_file)
        self._load_trades(trades_file)
        log.info("Loaded %d users and trades for %d logins from files",
                 len(self._users), len(self._trades_by_login))

    def disconnect(self):
        pass

    def _load_users(self, path):
        header, rows = _read_rows(path)
        idx = _build_index(header, USER_ALIASES)
        missing = [f for f in ("login", "name") if f not in idx]
        if missing:
            raise ValueError(f"USERS_FILE missing required columns {missing}. "
                             f"Found headers: {header}")
        for row in rows:
            if not row or all(c == "" for c in row):
                continue
            def get(field, default=""):
                return row[idx[field]] if field in idx and idx[field] < len(row) else default
            consent_raw = get("consent", "1")  # if no consent column, treat file as pre-filtered
            self._users.append({
                "login": int(_to_float(get("login"))),
                "name": str(get("name")).strip(),
                "email": str(get("email")).strip(),
                "telegram_id": str(get("telegram_id")).strip(),
                "consent": _to_bool_int(consent_raw) if "consent" in idx else 1,
            })

    def _load_trades(self, path):
        header, rows = _read_rows(path)
        idx = _build_index(header, TRADE_ALIASES)
        required = ["login", "profit"]
        missing = [f for f in required if f not in idx]
        if missing:
            raise ValueError(f"TRADES_FILE missing required columns {missing}. "
                             f"Found headers: {header}")
        for row in rows:
            if not row or all(c == "" for c in row):
                continue
            def get(field, default=""):
                return row[idx[field]] if field in idx and idx[field] < len(row) else default
            login = int(_to_float(get("login")))
            trade = {
                "login": login,
                "ticket": get("ticket", ""),
                "symbol": str(get("symbol", "")).strip(),
                "type": _to_direction(get("type", "buy")),
                "volume": _to_float(get("volume", 0)),
                "open_price": _to_float(get("open_price", 0)),
                "close_price": _to_float(get("close_price", 0)),
                "sl": _to_float(get("sl", 0)),
                "tp": _to_float(get("tp", 0)),
                "sl_hit": _to_bool_int(get("sl_hit", 0)),
                "tp_hit": _to_bool_int(get("tp_hit", 0)),
                "open_time": _to_time(get("open_time", "")),
                "close_time": _to_time(get("close_time", "")),
                "profit": _to_float(get("profit", 0)),
            }
            self._trades_by_login.setdefault(login, []).append(trade)

    def get_users(self):
        # CONSENT GATE: only opted-in users leave this method.
        consented = [u for u in self._users if u.get("consent") == 1]
        skipped = len(self._users) - len(consented)
        if skipped:
            log.info("Consent gate: %d user(s) skipped (no opt-in)", skipped)
        return [{k: u[k] for k in ("login", "name", "email", "telegram_id")}
                for u in consented]

    def get_closed_deals(self, login, since):
        """
        Return this login's trades. `since` filtering is applied only if the
        trade has a parseable open_time; otherwise the trade is included (a
        manual export is usually already scoped to the period you exported).
        """
        trades = self._trades_by_login.get(login, [])
        if since is None:
            return trades
        # Compare naive-to-naive: our open_time strings are naive UTC, so drop
        # any tzinfo on `since` to avoid aware/naive comparison errors.
        since_naive = since.replace(tzinfo=None)
        out = []
        for t in trades:
            ot = t.get("open_time", "")
            if not ot:
                out.append(t)
                continue
            try:
                if datetime.strptime(ot, "%Y-%m-%d %H:%M:%S") >= since_naive:
                    out.append(t)
            except ValueError:
                out.append(t)
        return out
